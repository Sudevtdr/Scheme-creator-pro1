import io
import re
import json
import pandas as pd
from flask import Flask, request, jsonify, render_template, send_file
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# --- Initialize Flask App ---
app = Flask(__name__)
app.config['JSON_SORT_KEYS'] = False
try: app.json.sort_keys = False
except AttributeError: pass

# --- Styling Data & Export Generators ---
themes = {
    "Default Blue": {"z": "#1f497d", "d": "#4f81bd", "s": "#dce6f1", "p": "#ffffff", "st": "#f2f2f2", "zt": "#e6e6e6", "gt": "#d4edda", "zf": "#ffffff", "df": "#ffffff", "sf": "#000000"},
    "Night Mode": {"z": "#121212", "d": "#2d2d2d", "s": "#404040", "p": "#1e1e1e", "st": "#333333", "zt": "#222222", "gt": "#004d40", "zf": "#ffffff", "df": "#eeeeee", "sf": "#dddddd"},
    "Desert Sand": {"z": "#5d4037", "d": "#8d6e63", "s": "#d7ccc8", "p": "#fffdf7", "st": "#efebe9", "zt": "#e4e0df", "gt": "#c8e6c9", "zf": "#ffffff", "df": "#ffffff", "sf": "#000000"},
    "Oceanic": {"z": "#004d40", "d": "#00838f", "s": "#b2ebf2", "p": "#ffffff", "st": "#e0f7fa", "zt": "#b2ebf2", "gt": "#a5d6a7", "zf": "#ffffff", "df": "#ffffff", "sf": "#000000"},
    "High Contrast": {"z": "#000000", "d": "#333333", "s": "#cccccc", "p": "#ffffff", "st": "#aaaaaa", "zt": "#888888", "gt": "#ffff00", "zf": "#ffffff", "df": "#ffffff", "sf": "#000000"},
    "Forest Green": {"z": "#1b5e20", "d": "#2e7d32", "s": "#c8e6c9", "p": "#f1f8e9", "st": "#e8f5e9", "zt": "#c8e6c9", "gt": "#a5d6a7", "zf": "#ffffff", "df": "#ffffff", "sf": "#000000"},
    "Sunset Orange": {"z": "#e65100", "d": "#ef6c00", "s": "#ffe0b2", "p": "#fff8e1", "st": "#ffecb3", "zt": "#ffe0b2", "gt": "#ffcc80", "zf": "#ffffff", "df": "#ffffff", "sf": "#000000"},
    "Royal Purple": {"z": "#4a148c", "d": "#6a1b9a", "s": "#e1bee7", "p": "#f3e5f5", "st": "#f3e5f5", "zt": "#e1bee7", "gt": "#ce93d8", "zf": "#ffffff", "df": "#ffffff", "sf": "#000000"},
    "Crimson Red": {"z": "#b71c1c", "d": "#c62828", "s": "#ffcdd2", "p": "#ffebee", "st": "#ffebee", "zt": "#ffcdd2", "gt": "#ef9a9a", "zf": "#ffffff", "df": "#ffffff", "sf": "#000000"},
    "Steel Slate": {"z": "#263238", "d": "#455a64", "s": "#cfd8dc", "p": "#eceff1", "st": "#eceff1", "zt": "#cfd8dc", "gt": "#b0bec5", "zf": "#ffffff", "df": "#ffffff", "sf": "#000000"},
    "Cyberpunk Neon": {"z": "#0d0221", "d": "#240046", "s": "#ff007f", "p": "#000000", "st": "#11001c", "zt": "#240046", "gt": "#00f0ff", "zf": "#00f0ff", "df": "#ff007f", "sf": "#ffffff"}
}

def get_contrasting_text(hex_code):
    if hex_code.lower() in ["white", "black", "#ffffff", "#000000"]: return "black" if hex_code.lower() in ["white", "#ffffff"] else "white"
    h = str(hex_code).lstrip('#')
    if len(h) < 6: return "black"
    r, g, b = tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
    luminance = (0.299*r + 0.587*g + 0.114*b)/255
    return "black" if luminance > 0.5 else "white"

def generate_styled_excel(df, source_tab, theme_colors, output_mode="full"):
    output = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)
    
    def to_hex(c):
        if str(c).lower() == "black": return "000000"
        if str(c).lower() == "white": return "FFFFFF"
        return str(c).lstrip('#').upper()
    
    chunks = {"Full Scheme": df}
    if output_mode == "zone_sheets":
        chunks = {}; current_z = "Unzoned"; z_data = []
        for _, row in df.iterrows():
            val0 = str(row.iloc[0])
            if "🏢 ZONE:" in val0 or "STANDBY / RESERVE" in val0:
                if z_data: chunks[current_z] = pd.DataFrame(z_data, columns=df.columns)
                z_data = []
                m = re.search(r'ZONE:\s*(.*?)\s*\|', val0)
                current_z = m.group(1).strip() if m else "Reserve"
                current_z = re.sub(r'[\\/*?:\[\]]', '', current_z)[:30]
            z_data.append(row)
        if z_data: chunks[current_z] = pd.DataFrame(z_data, columns=df.columns)
        if not chunks: chunks = {"Full Scheme": df}

    for title, chunk_df in chunks.items():
        ws = wb.create_sheet(title=title if title else "Sheet")
        columns = chunk_df.columns.tolist()
        ws.append(columns)
        for i in range(1, len(columns)+1): ws.cell(row=1, column=i).font = Font(bold=True)
        
        if source_tab == "preview":
            ws.column_dimensions['A'].width = 60
            for i in range(2, len(columns)+1): ws.column_dimensions[chr(64+i)].width = 15
        elif source_tab == "deployed":
            ws.column_dimensions['A'].width = 50
            for i, w in enumerate([15, 15, 15, 20, 15], 2): ws.column_dimensions[chr(64+i)].width = w
        elif source_tab == "matrix":
            ws.column_dimensions['A'].width = 35
            for i in range(2, len(columns)+1): ws.column_dimensions[chr(64+i)].width = 45
        elif source_tab == "totals":
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            for i in range(3, len(columns)+1): ws.column_dimensions[chr(64+i)].width = 15

        for _, row in chunk_df.iterrows():
            vals = row.tolist()
            ws.append(vals)
            current_row = ws.max_row
            
            val0 = str(vals[0])
            tags = []

            if source_tab == "totals":
                level = str(vals[0])
                fill_color, font_color, is_bold = None, "000000", False
                
                if level == "SCHEME TOTAL":
                    fill_color, font_color, is_bold = to_hex(theme_colors['gt']), to_hex(get_contrasting_text(theme_colors['gt'])), True
                elif level == "ZONE":
                    fill_color, font_color, is_bold = to_hex(theme_colors['z']), to_hex(theme_colors['zf']), True
                elif level == "  ↳ DIVISION":
                    fill_color, font_color, is_bold = to_hex(theme_colors['d']), to_hex(theme_colors['df']), True
                elif level == "      ↳ SECTOR":
                    fill_color, font_color, is_bold = to_hex(theme_colors['s']), to_hex(theme_colors['sf']), True
                
                if fill_color:
                    for i in range(1, len(columns) + 1):
                        cell = ws.cell(row=current_row, column=i)
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        cell.font = Font(color=font_color, bold=is_bold)
                continue

            if "🏢 ZONE:" in val0 or "STANDBY / RESERVE" in val0: tags.append('zone_header')
            elif "🛡️ DIV:" in val0: tags.append('div_header')
            elif "🎯 SEC:" in val0: tags.append('sec_header')
            elif "▼ DUTY POINTS" in val0: tags.append('point_header')
            elif "📍" in val0: tags.append('point_row')
            elif "👤" in val0: tags.append('person_row')
            elif "🌟 GRAND SCHEME" in val0: tags.append('grand_total_row')
            elif "∑ TOTAL" in val0 and "ZONE" in val0: tags.append('zone_total_row')
            elif "∑ TOTAL" in val0: tags.append('sec_total_row')
            elif "📜 SCHEME:" in val0: tags.append('main_heading_row')
            elif "📅 DATE:" in val0: tags.append('date_heading_row')
            else: tags.append('point_row')

            fill_color, font_color, is_bold = None, "000000", False
            if 'zone_header' in tags: fill_color, font_color, is_bold = to_hex(theme_colors['z']), to_hex(theme_colors['zf']), True
            elif 'div_header' in tags: fill_color, font_color, is_bold = to_hex(theme_colors['d']), to_hex(theme_colors['df']), True
            elif 'sec_header' in tags: fill_color, font_color, is_bold = to_hex(theme_colors['s']), to_hex(theme_colors['sf']), True
            elif 'grand_total_row' in tags: fill_color, font_color, is_bold = to_hex(theme_colors['gt']), to_hex(get_contrasting_text(theme_colors['gt'])), True
            elif 'zone_total_row' in tags: fill_color, font_color, is_bold = to_hex(theme_colors['zt']), to_hex(get_contrasting_text(theme_colors['zt'])), True
            elif 'sec_total_row' in tags: fill_color, font_color, is_bold = to_hex(theme_colors['st']), to_hex(get_contrasting_text(theme_colors['st'])), True
            elif 'main_heading_row' in tags: fill_color, font_color, is_bold = to_hex(theme_colors['z']), to_hex(theme_colors['zf']), True
            elif 'date_heading_row' in tags: fill_color, font_color, is_bold = to_hex(theme_colors['d']), to_hex(theme_colors['df']), True
            elif 'point_header' in tags:
                fill_color, font_color, is_bold = to_hex(theme_colors['st']), to_hex(get_contrasting_text(theme_colors['st'])), True
                ws.cell(row=current_row, column=1).font = Font(color=font_color, bold=True, underline="single")
                for c in range(2, len(columns)+1): ws.cell(row=current_row, column=c).font = Font(color=font_color, bold=True)
                for i in range(1, len(columns) + 1): ws.cell(row=current_row, column=i).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                continue
            elif 'person_row' in tags or 'point_row' in tags:
                fill_color, font_color, is_bold = to_hex(theme_colors['p']), to_hex(get_contrasting_text(theme_colors['p'])), False

            if fill_color:
                for i in range(1, len(columns) + 1):
                    cell = ws.cell(row=current_row, column=i)
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell.font = Font(color=font_color, bold=is_bold)
            elif source_tab == "matrix":
                 for i in range(1, len(columns) + 1): ws.cell(row=current_row, column=i).font = Font(color=font_color, bold=is_bold)

            if ('zone_header' in tags or 'div_header' in tags or 'sec_header' in tags) and source_tab in ["preview", "deployed"]:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columns))
            
            if 'main_heading_row' in tags or 'date_heading_row' in tags:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columns))
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                if 'main_heading_row' in tags: ws.cell(row=current_row, column=1).font = Font(color=font_color, bold=is_bold, size=14)
                if 'date_heading_row' in tags: ws.cell(row=current_row, column=1).font = Font(color=font_color, bold=is_bold, size=12)
                
            if source_tab == "matrix":
                for i in range(1, len(columns) + 1): ws.cell(row=current_row, column=i).alignment = Alignment(wrap_text=True, vertical='center')

    wb.save(output)
    return output.getvalue()

def generate_html_report(df, title, source_tab, theme_colors):
    z_bg, z_fg, d_bg, d_fg = theme_colors['z'], theme_colors['zf'], theme_colors['d'], theme_colors['df']
    s_bg, s_fg, p_bg, p_fg = theme_colors['s'], theme_colors['sf'], theme_colors['p'], get_contrasting_text(theme_colors['p'])
    st_bg, st_fg = theme_colors['st'], get_contrasting_text(theme_colors['st'])
    
    html = f"<html><head><meta charset='utf-8'><title>{title}</title>"
    html += f"<style>body {{ font-family: 'Segoe UI', sans-serif; margin: 20px; font-size: 12px; background-color: {p_bg}; color: {p_fg}; }} table {{ width: 100%; border-collapse: collapse; }} th, td {{ border: 1px solid #dddddd; padding: 6px; text-align: left; vertical-align: top; }} th {{ background-color: #f2f2f2; color: #000; }} .zone_header {{ background-color: {z_bg} !important; color: {z_fg} !important; font-weight: bold; }} .div_header {{ background-color: {d_bg} !important; color: {d_fg} !important; font-weight: bold; }} .sec_header {{ background-color: {s_bg} !important; color: {s_fg} !important; font-weight: bold; }} .point_header {{ background-color: {st_bg} !important; color: {st_fg} !important; font-weight: bold; text-decoration: underline; }} .person_row {{ background-color: {p_bg} !important; color: {p_fg} !important; }} .main_heading_row {{ background-color: {z_bg} !important; color: {z_fg} !important; font-weight: bold; font-size: 16px; text-align: center; }} .date_heading_row {{ background-color: {d_bg} !important; color: {d_fg} !important; font-weight: bold; font-size: 14px; text-align: center; }} @media print {{ tr td {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }} }}</style></head><body><h2>{title}</h2><table><thead><tr>"
    columns = df.columns.tolist()
    for col in columns: html += f"<th>{col}</th>"
    html += "</tr></thead><tbody>"
    for _, row in df.iterrows():
        vals = row.tolist()
        val0 = str(vals[0])
        if "🏢 ZONE" in val0 or "STANDBY" in val0: r_class = 'zone_header'
        elif "🛡️ DIV" in val0: r_class = 'div_header'
        elif "🎯 SEC" in val0: r_class = 'sec_header'
        elif "▼ DUTY POINTS" in val0: r_class = 'point_header'
        elif "📍" in val0: r_class = 'point_row'
        elif "👤" in val0: r_class = 'person_row'
        elif "🌟 GRAND SCHEME" in val0: r_class = 'grand_total_row'
        elif "∑ TOTAL" in val0 and "ZONE" in val0: r_class = 'zone_total_row'
        elif "∑ TOTAL" in val0: r_class = 'sec_total_row'
        elif "📜 SCHEME" in val0: r_class = 'main_heading_row'
        elif "📅 DATE" in val0: r_class = 'date_heading_row'
        else: r_class = 'point_row'
        if r_class in ['zone_header', 'div_header', 'sec_header'] and source_tab in ['preview', 'deployed']: html += f"<tr class='{r_class}'><td colspan='{len(columns)}'>{str(vals[0])}</td></tr>"
        elif r_class in ['main_heading_row', 'date_heading_row']: html += f"<tr class='{r_class}'><td colspan='{len(columns)}' style='text-align: center;'>{str(vals[0])}</td></tr>"
        else:
            html += f"<tr class='{r_class}'>"
            for val in vals: html += f"<td>{str(val).replace(chr(10), '<br>')}</td>"
            html += "</tr>"
    html += "</tbody></table></body></html>"
    return html.encode('utf-8')

# --- Helper Functions ---
def safe_int(val):
    try:
        return int(float(str(val).strip())) if str(val).strip() else 0
    except (ValueError, TypeError):
        return 0

def natural_key(text):
    return tuple(int(c) if c.isdigit() else c for c in re.split(r'(\d+)', str(text).strip().lower()))

def build_loc_str(*args):
    return ", ".join([str(x).strip() for x in args if str(x).strip()])

def parse_turn(text):
    m = re.search(r'(.*)\[(.*?)\]$', str(text).strip())
    if m: return m.group(1).strip().upper(), m.group(2).strip().upper()
    return str(text).strip().upper(), "GENERAL"

def get_compressed_person(p):
    parts = [p['name'], p['rank']]
    if p['gl']: parts.append(f"GL-{p['gl']}")
    if p['pen']: parts.append(f"PEN-{p['pen']}")
    if p['unit']: parts.append(p['unit'])
    if p['mob']: parts.append(f"MOB-{p['mob']}")
    return ", ".join(parts)

def to_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
    processed_data = output.getvalue()
    return processed_data

# --- Core Business Logic ---
def apply_chain_of_command(df):
    data = df.values.tolist()
    seen_zones, seen_divisions, seen_sectors = set(), set(), set()
    for r_idx, row in enumerate(data):
        while len(row) < len(df.columns): row.append("")
        for c in [4, 5, 6, 7, 8, 9]:
            if c < len(row):
                val = str(data[r_idx][c]).strip()
                if val:
                    try: data[r_idx][c] = str(int(float(val)))
                    except ValueError: data[r_idx][c] = ""
        zone, division, sector, point = str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip()
        
        # Auto-populate SI/ASI on dedicated header rows if no commander is provided
        if zone and not division and not sector and not point:
            if not any([data[r_idx][c] for c in [4, 5, 6, 7]]): data[r_idx][4] = "1"
        if division and not sector and not point:
            if not any([data[r_idx][c] for c in [5, 6, 7]]): data[r_idx][7] = "1"
        if sector and not point:
            if not any([data[r_idx][c] for c in [6, 7]]): data[r_idx][7] = "1"
            
        if zone:
            if zone not in seen_zones:
                seen_zones.add(zone)
            else:
                data[r_idx][4] = ""
        if division:
            if (zone, division) not in seen_divisions:
                seen_divisions.add((zone, division))
            else:
                data[r_idx][5] = ""
        if sector:
            if (zone, division, sector) not in seen_sectors:
                seen_sectors.add((zone, division, sector))
            else:
                data[r_idx][6] = ""
    return pd.DataFrame(data, columns=df.columns)

def get_aggregated_hierarchy(df):
    hierarchy = {}
    last_z = last_d = last_s = ""
    for _, row in df.iterrows():
        row = [str(x) for x in row.tolist()]
        while len(row) < 10: row.append("")
        z, d, s, p = row[0].strip(), row[1].strip(), row[2].strip(), row[3].strip()
        if z: last_z = z; last_d = ""; last_s = ""
        else: z = last_z
        if d: last_d = d; last_s = ""
        else: d = last_d
        if s: last_s = s
        else: s = last_s
        sp, dysp, ip = row[4].strip(), row[5].strip(), row[6].strip()
        si, cpo, wcpo = row[7].strip(), row[8].strip(), row[9].strip()
        z_key = z if z else "UNZONED"
        if not z and not d and not s: continue
        
        if z_key not in hierarchy: hierarchy[z_key] = {'sp': sp, 'dysp': '', 'ip': '', 'si': '', 'divs': {}}
        if sp and sp != "0": hierarchy[z_key]['sp'] = sp 
        if not d and not s and not p:
            if dysp and dysp != "0": hierarchy[z_key]['dysp'] = dysp
            if ip and ip != "0": hierarchy[z_key]['ip'] = ip
            if si and si != "0": hierarchy[z_key]['si'] = si

        if d:
            if d not in hierarchy[z_key]['divs']: hierarchy[z_key]['divs'][d] = {'dysp': dysp, 'ip': '', 'si': '', 'secs': {}}
            if dysp and dysp != "0": hierarchy[z_key]['divs'][d]['dysp'] = dysp
            if not s and not p:
                if ip and ip != "0": hierarchy[z_key]['divs'][d]['ip'] = ip
                if si and si != "0": hierarchy[z_key]['divs'][d]['si'] = si

            if s:
                if s not in hierarchy[z_key]['divs'][d]['secs']: hierarchy[z_key]['divs'][d]['secs'][s] = {'ip': ip, 'si': '', 'points': []}
                if ip and ip != "0": hierarchy[z_key]['divs'][d]['secs'][s]['ip'] = ip
                if not p:
                    if si and si != "0": hierarchy[z_key]['divs'][d]['secs'][s]['si'] = si
                
                if p: 
                    hierarchy[z_key]['divs'][d]['secs'][s]['points'].append((p, si, cpo, wcpo))
    return hierarchy

def build_readable_scheme_df(hierarchy, cmd_names, force_names, heading="", date=""):
    hier_data = []
    grand_si = grand_cpo = grand_wcpo = 0
    total_sp = total_dysp = total_ip = 0
    
    if heading: hier_data.append([f"📜 SCHEME: {heading.upper()}"] + [""] * len(force_names))
    if date: hier_data.append([f"📅 DATE: {date}"] + [""] * len(force_names))

    def build_oic_str(fallbacks):
        for val, name in fallbacks:
            if val and str(val) != "0": return f"({name.upper()}): {val}"
        # Fallback strictly to SI/ASI with a 0 count if completely empty
        return f"({fallbacks[-1][1].upper()}): 0"
        
    for z, z_data in hierarchy.items():
        oic_z = build_oic_str([(z_data.get('sp'), cmd_names[0]), (z_data.get('dysp'), cmd_names[1]), (z_data.get('ip'), cmd_names[2]), (z_data.get('si'), force_names[0])])
        if z != "UNZONED":
            hier_data.append([f"🏢 ZONE: {z.upper()} | OFFICER IN CHARGE {oic_z}", "", "", ""])
            total_sp += safe_int(z_data.get('sp'))
            total_dysp += safe_int(z_data.get('dysp'))
            total_ip += safe_int(z_data.get('ip'))
        zone_si = zone_cpo = zone_wcpo = 0
        if z != "UNZONED":
            zone_si += safe_int(z_data.get('si'))
            
        for d, d_data in z_data['divs'].items():
            oic_d = build_oic_str([(d_data.get('dysp'), cmd_names[1]), (d_data.get('ip'), cmd_names[2]), (d_data.get('si'), force_names[0])])
            hier_data.append([f"🛡️ DIV: {d.upper()} | OFFICER IN CHARGE {oic_d}", "", "", ""])
            total_dysp += safe_int(d_data.get('dysp'))
            total_ip += safe_int(d_data.get('ip'))
            zone_si += safe_int(d_data.get('si'))
            
            for s, s_data in d_data['secs'].items():
                oic_s = build_oic_str([(s_data.get('ip'), cmd_names[2]), (s_data.get('si'), force_names[0])])
                hier_data.append([f"🎯 SEC: {s.upper()} | OFFICER IN CHARGE {oic_s}", "", "", ""])
                total_ip += safe_int(s_data.get('ip'))
                hier_data.append(["        ▼ DUTY POINTS", force_names[0].upper(), force_names[1].upper(), force_names[2].upper()])
                point_counter = 1
                sec_si = safe_int(s_data.get('si'))
                sec_cpo = sec_wcpo = 0
                for pt in s_data['points']:
                    p_name = pt[0].upper() if pt[0] else "UNNAMED POINT"
                    hier_data.append([f"        {point_counter}. 📍 {p_name}", pt[1], pt[2], pt[3]])
                    sec_si += safe_int(pt[1]); sec_cpo += safe_int(pt[2]); sec_wcpo += safe_int(pt[3])
                    point_counter += 1
                hier_data.append([f"        ∑ TOTAL FOR {s.upper()}", str(sec_si), str(sec_cpo), str(sec_wcpo)])
                zone_si += sec_si; zone_cpo += sec_cpo; zone_wcpo += sec_wcpo
        if z != "UNZONED":
            hier_data.append([f"    ∑ TOTAL FOR {z.upper()}", str(zone_si), str(zone_cpo), str(zone_wcpo)])
            hier_data.append(["", "", "", ""])
        grand_si += zone_si; grand_cpo += zone_cpo; grand_wcpo += zone_wcpo
    if hierarchy:
        grand_txt = f"🌟 GRAND SCHEME TOTAL (CMDRS: {cmd_names[0].upper()}:{total_sp} | {cmd_names[1].upper()}:{total_dysp} | {cmd_names[2].upper()}:{total_ip})"
        hier_data.append([grand_txt, str(grand_si), str(grand_cpo), str(grand_wcpo)])

    # Post-process to hide empty commander rows as requested
    empty_indicator = f"OFFICER IN CHARGE ({force_names[0].upper()}): 0"
    for i in range(len(hier_data)):
        if empty_indicator in str(hier_data[i][0]):
            hier_data[i] = [""] * len(hier_data[i])

    return pd.DataFrame(hier_data, columns=["HIERARCHY / LOCATION"] + force_names)

def get_scheme_requirements(df, cmd_names, force_names):
    reqs = {}
    seen_z, seen_d, seen_s = set(), set(), set()
    last_z = last_d = last_s = ""
    def add_req(duty_str, rank, count):
        if not duty_str or count <= 0: return
        if duty_str not in reqs: reqs[duty_str] = {}
        reqs[duty_str][rank] = reqs[duty_str].get(rank, 0) + count
        
    for _, row in df.iterrows():
        row = [str(x) for x in row.tolist()]
        while len(row) < 10: row.append("")
        z, d, s, p = row[0].strip(), row[1].strip(), row[2].strip(), row[3].strip()
        if z: last_z = z; last_d = ""; last_s = ""
        else: z = last_z
        if d: last_d = d; last_s = ""
        else: d = last_d
        if s: last_s = s
        else: s = last_s
        
        sp, dysp, ip = safe_int(row[4]), safe_int(row[5]), safe_int(row[6])
        f1, f2, f3 = safe_int(row[7]), safe_int(row[8]), safe_int(row[9])
        z_key = z if z else "UNZONED"
        loc_z = z
        loc_d = f"{z}, {d}".strip(", ")
        loc_s = f"{z}, {d}, {s}".strip(", ")
        loc_p = f"{z}, {d}, {s}, {p}".strip(", ")
        
        if z and z not in seen_z: add_req(loc_z, cmd_names[0], sp); seen_z.add(z)
        if d and (z_key, d) not in seen_d: add_req(loc_d, cmd_names[1], dysp); seen_d.add((z_key, d))
        if s and (z_key, d, s) not in seen_s: add_req(loc_s, cmd_names[2], ip); seen_s.add((z_key, d, s))
        if p:
            add_req(loc_p, force_names[0], f1); add_req(loc_p, force_names[1], f2); add_req(loc_p, force_names[2], f3)
        elif s:
            add_req(loc_s, force_names[0], f1); add_req(loc_s, force_names[1], f2); add_req(loc_s, force_names[2], f3)
        elif d:
            add_req(loc_d, force_names[0], f1); add_req(loc_d, force_names[1], f2); add_req(loc_d, force_names[2], f3)
            if ip: add_req(loc_d, cmd_names[2], ip)
        elif z:
            add_req(loc_z, force_names[0], f1); add_req(loc_z, force_names[1], f2); add_req(loc_z, force_names[2], f3)
            if dysp: add_req(loc_z, cmd_names[1], dysp)
            if ip: add_req(loc_z, cmd_names[2], ip)
    return reqs

def build_deployed_data(scheme_df, nom_df, cmd_names, force_names, heading="", date=""):
    assignments = {}
    rank_prio = {force_names[0]: 1, force_names[1]: 2, force_names[2]: 3, cmd_names[0]: 0, cmd_names[1]: 0, cmd_names[2]: 0}
    
    if nom_df.empty:
        return pd.DataFrame(columns=["HIERARCHY / LOCATION / NAME", "RANK", "GL NUMBER", "PEN", "UNIT", "MOBILE"]), pd.DataFrame(columns=["Info"])
        
    for _, row in nom_df.iterrows():
        duty = str(row.get('Duty Allocation', '')).strip()
        if duty:
            if duty not in assignments: assignments[duty] = []
            assignments[duty].append({
                'name': str(row.get('Name', '')).strip().upper(),
                'rank': (str(row.get('Rank (Raw)', '')).strip() or str(row.get('Preferred Rank', '')).strip()).upper(),
                'gl': str(row.get('GL Number', '')).replace('.0',''),
                'pen': str(row.get('PEN', '')).replace('.0',''),
                'unit': str(row.get('Unit', '')).upper(),
                'mob': str(row.get('Mobile', '')).replace('.0',''),
                'prio': rank_prio.get(str(row.get('Preferred Rank', '')).strip(), 99)
            })
            
    for k in assignments: assignments[k].sort(key=lambda x: x['prio'])

    def format_cmdr_vertical(assigned_list):
        if not assigned_list: return "(NOT ASSIGNED)"
        res = []
        for p in assigned_list:
            s = f"{p['name']} ({p['rank']})"
            if p['mob'] and str(p['mob']).strip():
                s += f" | 📞 {p['mob']}"
            res.append(s)
        return " , ".join(res)
        
    def get_dynamic_oic_label(assignments_list, fallbacks, hier_node):
        if assignments_list:
            ranks = []
            for p in assignments_list:
                if p['rank'] not in ranks: ranks.append(p['rank'])
            return f"({', '.join(ranks)})"
        
        for key, name in fallbacks:
            if hier_node.get(key): return f"({name.upper()})"
        return f"({fallbacks[-1][1].upper()})"

    unique_turns = set()
    for _, row in scheme_df.iterrows():
        for i in range(4):
            val = str(row.iloc[i]).strip()
            if val:
                _, t = parse_turn(val)
                if t != 'GENERAL': unique_turns.add(t)
    
    turn_list = sorted(list(unique_turns))[:4]
    if not turn_list: turn_list = ["TURN 1"]
    matrix_cols = ["HIERARCHY / LOCATION"] + turn_list

    dep_data = []
    matrix_data = []
    
    if heading: dep_data.append([f"📜 SCHEME: {heading.upper()}", "", "", "", "", ""]); matrix_data.append([f"📜 SCHEME: {heading.upper()}"] + [""] * (len(matrix_cols)-1))
    if date: dep_data.append([f"📅 DATE: {date}", "", "", "", "", ""]); matrix_data.append([f"📅 DATE: {date}"] + [""] * (len(matrix_cols)-1))

    hierarchy = get_aggregated_hierarchy(scheme_df)
    matrix_struct = {}

    for z, z_data in hierarchy.items():
        bz, tz = parse_turn(z)
        if bz not in matrix_struct: matrix_struct[bz] = {'cmdrs': {}, 'divs': {}}
        
        if z != "UNZONED":
            sp_str = build_loc_str(z)
            sp_assigned = assignments.get(sp_str, [])
            label = get_dynamic_oic_label(sp_assigned, [('sp', cmd_names[0]), ('dysp', cmd_names[1]), ('ip', cmd_names[2]), ('si', force_names[0])], z_data)
            if sp_str in assignments: matrix_struct[bz]['cmdrs'][tz] = sp_assigned
            dep_data.append([f"🏢 ZONE: {z.upper()} | OFFICER IN CHARGE {label}: {format_cmdr_vertical(sp_assigned)}", "", "", "", "", ""])
            
        for d, d_data in z_data['divs'].items():
            bd, td = parse_turn(d)
            if bd not in matrix_struct[bz]['divs']: matrix_struct[bz]['divs'][bd] = {'cmdrs': {}, 'secs': {}}
            
            dysp_str = build_loc_str(z, d) if z != "UNZONED" else build_loc_str(d)
            dysp_assigned = assignments.get(dysp_str, [])
            label = get_dynamic_oic_label(dysp_assigned, [('dysp', cmd_names[1]), ('ip', cmd_names[2]), ('si', force_names[0])], d_data)
            if dysp_str in assignments: matrix_struct[bz]['divs'][bd]['cmdrs'][td] = dysp_assigned
            dep_data.append([f"🛡️ DIV: {d.upper()} | OFFICER IN CHARGE {label}: {format_cmdr_vertical(dysp_assigned)}", "", "", "", "", ""])
            
            for s, s_data in d_data['secs'].items():
                bs, ts = parse_turn(s)
                if bs not in matrix_struct[bz]['divs'][bd]['secs']: matrix_struct[bz]['divs'][bd]['secs'][bs] = {'cmdrs': {}, 'points': {}}
                
                ip_str = build_loc_str(z, d, s) if z != "UNZONED" else build_loc_str(d, s)
                ip_assigned = assignments.get(ip_str, [])
                label = get_dynamic_oic_label(ip_assigned, [('ip', cmd_names[2]), ('si', force_names[0])], s_data)
                if ip_str in assignments: matrix_struct[bz]['divs'][bd]['secs'][bs]['cmdrs'][ts] = ip_assigned
                dep_data.append([f"🎯 SEC: {s.upper()} | OFFICER IN CHARGE {label}: {format_cmdr_vertical(ip_assigned)}", "", "", "", "", ""])
                
                point_counter = 1
                for pt in s_data['points']:
                    p_name = pt[0].upper() if pt[0] else "UNNAMED POINT"
                    bp, tp = parse_turn(p_name)
                    if bp not in matrix_struct[bz]['divs'][bd]['secs'][bs]['points']: matrix_struct[bz]['divs'][bd]['secs'][bs]['points'][bp] = {}
                    
                    dep_data.append([f"    {point_counter}. 📍 {p_name}", "RANK", "GL NUMBER", "PEN", "UNIT", "MOBILE"])
                    
                    p_str = build_loc_str(z, d, s, pt[0]) if z != "UNZONED" else build_loc_str(d, s, pt[0])
                    if p_str in assignments:
                        matrix_struct[bz]['divs'][bd]['secs'][bs]['points'][bp][tp] = assignments[p_str]
                        for person in assignments[p_str]:
                            dep_data.append([f"          👤 {person['name']}", person['rank'], person['gl'], person['pen'], person['unit'], person['mob']])
                    else:
                        dep_data.append([f"          ⚠️ (NO PERSONNEL ASSIGNED YET)", "", "", "", "", ""])
                    point_counter += 1
            
    if "Standby / Reserve" in assignments:
        res_assigned = assignments["Standby / Reserve"]
        dep_data.append(["", "", "", "", "", ""])
        dep_data.append(["🚨 STANDBY / RESERVE FORCE", "RANK", "GL", "PEN", "UNIT", "MOBILE"])
        for person in res_assigned:
            dep_data.append([f"          👤 {person['name']}", person['rank'], person['gl'], person['pen'], person['unit'], person['mob']])

    def insert_matrix_block(name_prefix, data_dict):
        if not data_dict:
            matrix_data.append([name_prefix, "(NONE)"] + [""]*(len(turn_list)-1))
            return
            
        max_p = max((len(lst) for lst in data_dict.values()), default=0)
        for i in range(max_p):
            row_vals = [name_prefix if i == 0 else ""]
            for t in turn_list:
                if t in data_dict and i < len(data_dict[t]): row_vals.append(get_compressed_person(data_dict[t][i]))
                elif 'GENERAL' in data_dict and i < len(data_dict['GENERAL']): row_vals.append(get_compressed_person(data_dict['GENERAL'][i]))
                else: row_vals.append("")
            matrix_data.append(row_vals)

    for bz, z_data in matrix_struct.items():
        if bz != "UNZONED": insert_matrix_block(f"🏢 ZONE: {bz.upper()} | OFFICER IN CHARGE", z_data['cmdrs'])
        for bd, d_data in z_data['divs'].items():
            insert_matrix_block(f"🛡️ DIV: {bd.upper()} | OFFICER IN CHARGE", d_data['cmdrs'])
            for bs, s_data in d_data['secs'].items():
                insert_matrix_block(f"🎯 SEC: {bs.upper()} | OFFICER IN CHARGE", s_data['cmdrs'])
                pt_counter = 1
                for bp, p_data in s_data['points'].items():
                    insert_matrix_block(f"    {pt_counter}. 📍 {bp.upper()}", p_data)
                    pt_counter += 1

    # Post-process to hide empty commander rows in deployment and matrix sheets
    empty_indicator_dep = f"OFFICER IN CHARGE ({force_names[0].upper()}): (NOT ASSIGNED)"
    for i in range(len(dep_data)):
        if empty_indicator_dep in str(dep_data[i][0]):
            dep_data[i] = [""] * len(dep_data[i])
            
    for i in range(len(matrix_data)):
        if "OFFICER IN CHARGE" in str(matrix_data[i][0]) and len(matrix_data[i]) > 1 and "(NONE)" in str(matrix_data[i][1]):
            matrix_data[i] = [""] * len(matrix_data[i])
            
    return pd.DataFrame(dep_data, columns=["HIERARCHY / LOCATION / NAME", "RANK", "GL NUMBER", "PEN", "UNIT", "MOBILE"]), pd.DataFrame(matrix_data, columns=matrix_cols)

def get_manpower_totals_df(df, cmd_names, force_names):
    grand_total = [0]*6; zone_totals = {}; div_totals = {}; sec_totals = {}
    last_z = last_d = last_s = ""
    all_force_cols = cmd_names + force_names
    
    ordered_zones = []
    ordered_divs_by_zone = {}
    ordered_secs_by_div = {}

    for _, row in df.iterrows():
        z, d, s = str(row.get('Zone', '')).strip(), str(row.get('Division', '')).strip(), str(row.get('Sector', '')).strip()
        if z: last_z = z; last_d = ""; last_s = ""
        else: z = last_z
        if d: last_d = d; last_s = ""
        else: d = last_d
        if s: last_s = s
        else: s = last_s
        
        if not z and not d and not s: continue
        z_key = z if z else "UNZONED"
        forces = [safe_int(row.get(col, 0)) for col in all_force_cols]
        
        if z_key not in zone_totals:
            zone_totals[z_key] = [0]*6
            ordered_zones.append(z_key)
            
        if d:
            div_key = (z_key, d)
            if div_key not in div_totals:
                div_totals[div_key] = [0]*6
                if z_key not in ordered_divs_by_zone:
                    ordered_divs_by_zone[z_key] = []
                ordered_divs_by_zone[z_key].append(d)
                
        if s:
            sec_key = (z_key, d, s)
            if sec_key not in sec_totals:
                sec_totals[sec_key] = [0]*6
                div_sec_key = (z_key, d)
                if div_sec_key not in ordered_secs_by_div:
                    ordered_secs_by_div[div_sec_key] = []
                ordered_secs_by_div[div_sec_key].append(s)
        
        for i in range(6): 
            grand_total[i] += forces[i]
            zone_totals[z_key][i] += forces[i]
            if d: div_totals[(z_key, d)][i] += forces[i]
            if s: sec_totals[(z_key, d, s)][i] += forces[i]

    totals_data = []
    for z_key in ordered_zones:
        for d_key in ordered_divs_by_zone.get(z_key, []):
            for s_key in ordered_secs_by_div.get((z_key, d_key), []):
                totals_data.append(["      ↳ SECTOR", s_key] + sec_totals[(z_key, d_key, s_key)])
            totals_data.append(["  ↳ DIVISION", d_key] + div_totals[(z_key, d_key)])
        totals_data.append(["ZONE", z_key] + zone_totals[z_key])
        
    totals_data.append(["SCHEME TOTAL", "All Points"] + grand_total)
                        
    return pd.DataFrame(totals_data, columns=["Level", "Name"] + cmd_names + force_names)


# --- Basic Web API Endpoints ---

@app.route('/')
def index():
    return render_template('index.html')

def align_with_select(data, df):
    cmd_names = data.get('cmd_names', ["SP", "DySP", "IP"])
    force_names = data.get('force_names', ["SI/ASI", "SCPO/CPO", "WSCPO/WCPO"])
    expected_cols = ["Zone", "Division", "Sector", "Point"] + cmd_names + force_names
    
    selects = df['Select'] if 'Select' in df.columns else [False] * len(df)
    for c in expected_cols:
        if c not in df.columns: df[c] = ""
    df = df[expected_cols]
    
    aligned_df = apply_chain_of_command(df)
    aligned_df['Select'] = list(selects)
    return aligned_df

@app.route('/api/align-scheme', methods=['POST'])
def align_scheme():
    data = request.json
    df = pd.DataFrame(data['scheme_data']).fillna("")
    aligned_df = align_with_select(data, df)
    return jsonify(aligned_df.to_dict('records'))

@app.route('/api/sort-scheme', methods=['POST'])
def sort_scheme():
    data = request.json
    df = pd.DataFrame(data['scheme_data']).fillna("")
    df['sort_key_z'] = df.get('Zone', '').apply(lambda x: natural_key(str(x)))
    df['sort_key_d'] = df.get('Division', '').apply(lambda x: natural_key(str(x)))
    df = df.sort_values(by=['sort_key_z', 'sort_key_d']).drop(columns=['sort_key_z', 'sort_key_d']).reset_index(drop=True)
    aligned_df = align_with_select(data, df)
    return jsonify(aligned_df.to_dict('records'))

@app.route('/api/upload-scheme', methods=['POST'])
def upload_scheme():
    if 'file' not in request.files: return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    cmd_names = request.form.get('cmd_names', 'SP,DySP,IP').split(',')
    force_names = request.form.get('force_names', 'SI/ASI,SCPO/CPO,WSCPO/WCPO').split(',')
    headers = ["Zone", "Division", "Sector", "Point"] + cmd_names + force_names
    try: df = pd.read_excel(file).fillna("").astype(str)
    except Exception as e: return jsonify({'error': str(e)}), 400
    data_list = df.values.tolist()
    expected_len = len(headers)
    cleaned = []
    for row in data_list:
        row = list(row)
        while len(row) < expected_len: row.append("")
        cleaned.append(row[:expected_len])
    df_clean = pd.DataFrame(cleaned, columns=headers)
    df_clean["Select"] = False
    aligned_df = apply_chain_of_command(df_clean)
    aligned_df["Select"] = False
    return jsonify(aligned_df.to_dict('records'))

@app.route('/api/upload-nominal', methods=['POST'])
def upload_nominal():
    if 'file' not in request.files: return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    try:
        df = pd.read_excel(file).fillna("")
        nom_headers = ["Sl No", "Name", "Rank (Raw)", "GL Number", "PEN", "Unit", "Mobile", "Remarks", "Preferred Rank", "Duty Allocation"]
        df = df.iloc[:, :len(nom_headers)]
        df.columns = nom_headers[:len(df.columns)]
        for missing in nom_headers:
            if missing not in df.columns: df[missing] = ""
        mask = df.astype(str).apply(lambda s: s.str.strip()).ne("").any(axis=1)
        df = df[mask].astype(str)
        if 'Assignment Type' not in df.columns: df['Assignment Type'] = ""
        
        existing_data_str = request.form.get('existing_data', '[]')
        if existing_data_str and existing_data_str.strip() != '[]':
            existing_df = pd.DataFrame(json.loads(existing_data_str)).fillna("")
            if not existing_df.empty:
                df = pd.concat([existing_df, df], ignore_index=True).fillna("")
                
        return jsonify(df.to_dict('records'))
    except Exception as e: return jsonify({'error': str(e)}), 400

@app.route('/api/clean-ranks', methods=['POST'])
def clean_ranks():
    data = request.json
    df = pd.DataFrame(data['nom_data']).fillna("")
    cmd_names = data.get('cmd_names', ["SP", "DySP", "IP"])
    force_names = data.get('force_names', ["SI/ASI", "SCPO/CPO", "WSCPO/WCPO"])
    if 'Preferred Rank' not in df.columns: df['Preferred Rank'] = ""
    for idx, row in df.iterrows():
        raw = str(row.get('Rank (Raw)', '')).upper().replace(".", "").replace(" ", "").strip()
        pref = ""
        if raw in ["IP", "IOP", "ISHO", "CI", "INSPECTOR", "API", "RI", "DI"]: pref = cmd_names[2]
        elif raw in ["DYSP", "AC", "DSP", "ASP", "DC", "ADSP"]: pref = cmd_names[1]
        elif raw in ["SP"]: pref = cmd_names[0]
        elif raw in ["SI", "DSI", "ASI", "JRSI", "GSI", "WSI", "APSI", "APASI", "RSI", "RASI", "GRASI", "GRSI", "GASI", "SI(G)", "ASI(G)", "SGT"]: pref = force_names[0]
        elif raw in ["PC", "GSCPO", "CPO", "SCPO(G)", "HAV", "HDR", "HC","RTPC", "PCTELE", "SCPO", "HG"]: pref = force_names[1]
        elif "W" in raw: pref = force_names[2]
        if pref: df.at[idx, 'Preferred Rank'] = pref
    return jsonify(df.to_dict('records'))

@app.route('/api/auto-allocate', methods=['POST'])
def auto_allocate():
    data = request.json
    nom_df = pd.DataFrame(data['nom_data']).fillna("")
    scheme_df = pd.DataFrame(data['scheme_data']).fillna("")
    cmd_names = data.get('cmd_names', ["SP", "DySP", "IP"])
    force_names = data.get('force_names', ["SI/ASI", "SCPO/CPO", "WSCPO/WCPO"])
    if 'Duty Allocation' not in nom_df.columns: nom_df['Duty Allocation'] = ""
    if 'Assignment Type' not in nom_df.columns: nom_df['Assignment Type'] = ""
    reqs = get_scheme_requirements(scheme_df, cmd_names, force_names)
    import copy; working_reqs = copy.deepcopy(reqs)
    for idx, row in nom_df.iterrows():
        r, d = str(row.get('Preferred Rank', '')).strip(), str(row.get('Duty Allocation', '')).strip()
        if d and d != "Standby / Reserve" and d in working_reqs and r in working_reqs[d]:
            if working_reqs[d][r] > 0: working_reqs[d][r] -= 1
    for idx, row in nom_df.iterrows():
        r, d = str(row.get('Preferred Rank', '')).strip(), str(row.get('Duty Allocation', '')).strip()
        if not d and r:
            for pt, pt_req in working_reqs.items():
                if r in pt_req and pt_req[r] > 0:
                    nom_df.at[idx, 'Duty Allocation'] = pt; nom_df.at[idx, 'Assignment Type'] = "Auto"
                    pt_req[r] -= 1; break
    return jsonify(nom_df.to_dict('records'))

@app.route('/api/qrt-sweep', methods=['POST'])
def qrt_sweep():
    nom_df = pd.DataFrame(request.json['nom_data']).fillna("")
    if 'Duty Allocation' not in nom_df.columns: nom_df['Duty Allocation'] = ""
    if 'Assignment Type' not in nom_df.columns: nom_df['Assignment Type'] = ""
    for idx, row in nom_df.iterrows():
        if str(row.get('Preferred Rank', '')).strip() and not str(row.get('Duty Allocation', '')).strip():
            nom_df.at[idx, 'Duty Allocation'] = "Standby / Reserve"
            nom_df.at[idx, 'Assignment Type'] = "Auto"
    return jsonify(nom_df.to_dict('records'))

@app.route('/api/reset-auto', methods=['POST'])
def reset_auto():
    nom_df = pd.DataFrame(request.json['nom_data']).fillna("")
    if 'Duty Allocation' not in nom_df.columns: nom_df['Duty Allocation'] = ""
    if 'Assignment Type' not in nom_df.columns: nom_df['Assignment Type'] = ""
    for idx, row in nom_df.iterrows():
        if str(row.get('Assignment Type', '')).strip() == "Auto":
            nom_df.at[idx, 'Duty Allocation'] = ""
            nom_df.at[idx, 'Assignment Type'] = ""
    return jsonify(nom_df.to_dict('records'))

@app.route('/api/clone-rows', methods=['POST'])
def clone_rows():
    data = request.json
    df = pd.DataFrame(data['scheme_data']).fillna("")
    selected_indices = data.get('selected_indices', [])
    level = data.get('level', 'Point')
    labels = data.get('labels', ['Day', 'Night'])
    if not selected_indices or not labels: return jsonify(align_with_select(data, df).to_dict('records'))
    min_idx, max_idx = min(selected_indices), max(selected_indices)
    top_half = df.iloc[:min_idx]; bottom_half = df.iloc[max_idx+1:]; block = df.iloc[min_idx:max_idx+1]
    lvl_map = {"Zone": 0, "Division": 1, "Sector": 2, "Point": 3}
    c_lvl = lvl_map.get(level, 3)
    new_blocks = []
    for lbl in labels:
        suffix = f" [{lbl}]"
        new_block = block.copy()
        for idx in new_block.index:
            if idx in selected_indices:
                if c_lvl <= 0 and str(new_block.at[idx, 'Zone']).strip(): new_block.at[idx, 'Zone'] = str(new_block.at[idx, 'Zone']).strip() + suffix
                if c_lvl <= 1 and str(new_block.at[idx, 'Division']).strip(): new_block.at[idx, 'Division'] = str(new_block.at[idx, 'Division']).strip() + suffix
                if c_lvl <= 2 and str(new_block.at[idx, 'Sector']).strip(): new_block.at[idx, 'Sector'] = str(new_block.at[idx, 'Sector']).strip() + suffix
                if c_lvl <= 3:
                    p_val = str(new_block.at[idx, 'Point']).strip()
                    new_block.at[idx, 'Point'] = p_val + suffix if p_val else f"Point {suffix}"
            new_block.at[idx, 'Select'] = False
        new_blocks.append(new_block)
    new_df = pd.concat([top_half] + new_blocks + [bottom_half], ignore_index=True)
    return jsonify(align_with_select(data, new_df).to_dict('records'))

@app.route('/api/group-rows', methods=['POST'])
def group_rows():
    data = request.json
    df = pd.DataFrame(data['scheme_data']).fillna("")
    selected_indices = data.get('selected_indices', [])
    level = data.get('level')
    name = data.get('name', '')
    cmd_names = data.get('cmd_names', ["SP", "DySP", "IP"])
    if not selected_indices: return jsonify(align_with_select(data, df).to_dict('records'))
    min_idx, max_idx = min(selected_indices), max(selected_indices)
    if level == "Zone":
        df.loc[min_idx:max_idx, 'Zone'] = f"Zone - {name}" if not name.lower().startswith('zone') else name
        df.loc[min_idx:max_idx, ['Division', 'Sector'] + cmd_names] = ""
    elif level == "Division":
        df.loc[min_idx:max_idx, 'Division'] = f"Division - {name}" if not name.lower().startswith('div') else name
        df.loc[min_idx:max_idx, ['Sector'] + cmd_names[1:]] = ""
    elif level == "Sector":
        df.loc[min_idx:max_idx, 'Sector'] = f"Sector - {name}" if not name.lower().startswith('sec') else name
        df.loc[min_idx:max_idx, [cmd_names[2]]] = ""
    df['Select'] = False
    return jsonify(align_with_select(data, df).to_dict('records'))

@app.route('/api/duplicate-rows', methods=['POST'])
def duplicate_rows():
    data = request.json
    df = pd.DataFrame(data['scheme_data']).fillna("")
    selected_indices = data.get('selected_indices', [])
    copies = int(data.get('copies', 1))
    if not selected_indices: return jsonify(align_with_select(data, df).to_dict('records'))
    new_data = []
    for idx, row in df.iterrows():
        row_copy = row.copy()
        row_copy['Select'] = False
        new_data.append(row_copy)
        if idx in selected_indices:
            for _ in range(copies): new_data.append(row_copy)
    return jsonify(align_with_select(data, pd.DataFrame(new_data)).to_dict('records'))

@app.route('/api/fill-down', methods=['POST'])
def fill_down():
    data = request.json
    df = pd.DataFrame(data['scheme_data']).fillna("")
    selected_indices = data.get('selected_indices', [])
    cols_to_fill = data.get('columns', [])
    if not selected_indices or not cols_to_fill: return jsonify(align_with_select(data, df).to_dict('records'))
    min_idx, max_idx = min(selected_indices), max(selected_indices)
    for col in cols_to_fill:
        if col in df.columns:
            val = df.at[min_idx, col]
            df.loc[min_idx+1:max_idx, col] = val
    df['Select'] = False
    return jsonify(align_with_select(data, df).to_dict('records'))

@app.route('/api/manpower-totals', methods=['POST'])
def manpower_totals():
    data = request.json
    totals_df = get_manpower_totals_df(pd.DataFrame(data['scheme_data']).fillna(""), data['cmd_names'], data['force_names'])
    return jsonify(totals_df.to_dict('records'))

@app.route('/api/readable-scheme', methods=['POST'])
def readable_scheme():
    data = request.json
    hierarchy = get_aggregated_hierarchy(pd.DataFrame(data['scheme_data']).fillna(""))
    readable_df = build_readable_scheme_df(hierarchy, data['cmd_names'], data['force_names'], data.get('heading', ''), data.get('date', ''))
    return jsonify(readable_df.to_dict('records'))

@app.route('/api/deployed-data', methods=['POST'])
def deployed_data():
    data = request.json
    dep_df, mat_df = build_deployed_data(pd.DataFrame(data['scheme_data']).fillna(""), pd.DataFrame(data.get('nom_data', [])).fillna(""), data['cmd_names'], data['force_names'], data.get('heading', ''), data.get('date', ''))
    return jsonify({"deployed": dep_df.to_dict('records'), "matrix": mat_df.to_dict('records')})


# --- EXPORT / DOWNLOAD ENDPOINTS (From Streamlit Parity) ---

@app.route('/api/download/raw-scheme', methods=['POST'])
def download_raw_scheme():
    df = pd.DataFrame(request.json['scheme_data']).fillna("").drop(columns=["Select"], errors="ignore")
    return send_file(io.BytesIO(to_excel(df)), download_name="Raw_Scheme.xlsx", as_attachment=True)

@app.route('/api/download/totals', methods=['POST'])
def download_totals():
    data = request.json
    totals_df = get_manpower_totals_df(pd.DataFrame(data['scheme_data']).fillna(""), data['cmd_names'], data['force_names'])
    excel_data = generate_styled_excel(totals_df, "totals", themes[data.get('theme', 'Default Blue')], "full")
    return send_file(io.BytesIO(excel_data), download_name="Manpower_Totals.xlsx", as_attachment=True)

@app.route('/api/download/readable-excel', methods=['POST'])
def download_readable_excel():
    data = request.json
    readable_df = build_readable_scheme_df(get_aggregated_hierarchy(pd.DataFrame(data['scheme_data']).fillna("")), data['cmd_names'], data['force_names'], data.get('heading', ''), data.get('date', ''))
    excel_data = generate_styled_excel(readable_df, "preview", themes[data.get('theme', 'Default Blue')], "full")
    return send_file(io.BytesIO(excel_data), download_name="Readable_Scheme.xlsx", as_attachment=True)

@app.route('/api/download/readable-html', methods=['POST'])
def download_readable_html():
    data = request.json
    readable_df = build_readable_scheme_df(get_aggregated_hierarchy(pd.DataFrame(data['scheme_data']).fillna("")), data['cmd_names'], data['force_names'], data.get('heading', ''), data.get('date', ''))
    html_data = generate_html_report(readable_df, "Readable Scheme", "preview", themes[data.get('theme', 'Default Blue')])
    return send_file(io.BytesIO(html_data), download_name="Readable_Scheme.html", as_attachment=True)

@app.route('/api/download/nom-template', methods=['POST'])
def download_nom_template():
    nom_headers = ["Sl No", "Name", "Rank (Raw)", "GL Number", "PEN", "Unit", "Mobile", "Remarks"]
    return send_file(io.BytesIO(to_excel(pd.DataFrame(columns=nom_headers))), download_name="Blank_Roster_Template.xlsx", as_attachment=True)

@app.route('/api/download/nom-roll', methods=['POST'])
def download_nom_roll():
    df = pd.DataFrame(request.json['nom_data']).fillna("").drop(columns=["Assignment Type"], errors="ignore")
    mask = df.astype(str).apply(lambda s: s.str.strip()).ne("").any(axis=1)
    return send_file(io.BytesIO(to_excel(df[mask])), download_name="Nominal_Roll.xlsx", as_attachment=True)

@app.route('/api/download/deployed-excel', methods=['POST'])
def download_deployed_excel():
    data = request.json
    dep_df, _ = build_deployed_data(pd.DataFrame(data['scheme_data']).fillna(""), pd.DataFrame(data.get('nom_data', [])).fillna(""), data['cmd_names'], data['force_names'], data.get('heading', ''), data.get('date', ''))
    excel_data = generate_styled_excel(dep_df, "deployed", themes[data.get('theme', 'Default Blue')], data.get('mode', 'full'))
    name = "Deployed_Zones.xlsx" if data.get('mode') == 'zone_sheets' else "Deployed_Sheet.xlsx"
    return send_file(io.BytesIO(excel_data), download_name=name, as_attachment=True)

@app.route('/api/download/deployed-html', methods=['POST'])
def download_deployed_html():
    data = request.json
    dep_df, _ = build_deployed_data(pd.DataFrame(data['scheme_data']).fillna(""), pd.DataFrame(data.get('nom_data', [])).fillna(""), data['cmd_names'], data['force_names'], data.get('heading', ''), data.get('date', ''))
    html_data = generate_html_report(dep_df, "Deployed Scheme", "deployed", themes[data.get('theme', 'Default Blue')])
    return send_file(io.BytesIO(html_data), download_name="Deployed_Report.html", as_attachment=True)

@app.route('/api/download/matrix-excel', methods=['POST'])
def download_matrix_excel():
    data = request.json
    _, mat_df = build_deployed_data(pd.DataFrame(data['scheme_data']).fillna(""), pd.DataFrame(data.get('nom_data', [])).fillna(""), data['cmd_names'], data['force_names'], data.get('heading', ''), data.get('date', ''))
    excel_data = generate_styled_excel(mat_df, "matrix", themes[data.get('theme', 'Default Blue')], data.get('mode', 'full'))
    name = "Matrix_Zones.xlsx" if data.get('mode') == 'zone_sheets' else "Matrix_Sheet.xlsx"
    return send_file(io.BytesIO(excel_data), download_name=name, as_attachment=True)

@app.route('/api/download/matrix-html', methods=['POST'])
def download_matrix_html():
    data = request.json
    _, mat_df = build_deployed_data(pd.DataFrame(data['scheme_data']).fillna(""), pd.DataFrame(data.get('nom_data', [])).fillna(""), data['cmd_names'], data['force_names'], data.get('heading', ''), data.get('date', ''))
    html_data = generate_html_report(mat_df, "Matrix Matrix", "matrix", themes[data.get('theme', 'Default Blue')])
    return send_file(io.BytesIO(html_data), download_name="Matrix_Report.html", as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
